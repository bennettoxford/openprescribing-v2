import csv
import logging
import re
from datetime import datetime
from pathlib import Path
from tempfile import TemporaryDirectory
from zipfile import ZipFile

from bs4 import BeautifulSoup
from openpyxl import load_workbook

from openprescribing.data.utils.csv_to_parquet import csv_to_parquet
from openprescribing.data.utils.http_session import HTTPSession


log = logging.getLogger(__name__)


def fetch(directory):
    """Fetch most recent BNF SNOMED mapping file from
    https://www.nhsbsa.nhs.uk/prescription-data/understanding-our-data/bnf-snomed-mapping.

    As far as we're concerned, "SNOMED code" and "dm+d code" are interchangeable, so we
    just talk about dm+d.
    """

    dataset_dir = directory / "dmd_bnf_map"
    dataset_dir.mkdir(exist_ok=True, parents=True)

    http = HTTPSession(
        "https://www.nhsbsa.nhs.uk/",
        # We start with DEBUG level logging while we're checking if there's anything new
        # to fetch
        log=log.debug,
    )
    rsp = http.get("prescription-data/understanding-our-data/bnf-snomed-mapping")
    doc = BeautifulSoup(rsp.text, "html.parser")
    hrefs = [
        a["href"]
        for a in doc.find_all("a", href=True)
        if re.search(r"BNF%20Snomed%20Mapping%20data%20\d{8}.zip", a["href"])
    ]
    href = sorted(hrefs)[-1]
    a = get_single_item(a for a in doc.find_all("a", href=True) if a["href"] == href)
    release_month_year = re.search(r"\w+ \d{4}", a.text).group()
    release_date = datetime.strptime(release_month_year, "%B %Y").date()
    published_date = datetime.strptime(href.split(".")[0][-8:], "%Y%m%d").date()
    release_id = f"dmd_bnf_map_{release_date}_{published_date}"
    release_path = dataset_dir / f"{release_id}.parquet"

    if release_path.exists():
        log.debug(f"Already fetched a file for this release: {release_id}")
        return

    # Any requests we now make will be to fetch new files so we log at INFO level
    http.log = log.info

    with TemporaryDirectory() as tmp_dir:
        tmp_dir = Path(tmp_dir)
        zip_path = tmp_dir / "download.zip"
        csv_path = tmp_dir / "mapping.csv"

        rsp = http.download_to_file(href, zip_path)
        with ZipFile(zip_path) as zf:
            zf.extractall(tmp_dir)

        xlsx_path = get_single_item(tmp_dir.glob("*.xlsx"))
        wb = load_workbook(filename=xlsx_path)

        with open(csv_path, "w") as f:
            writer = csv.writer(f)
            writer.writerows(wb.active.iter_rows(values_only=True))

        csv_to_parquet(csv_path, release_path)


def get_single_item(iterable):
    items = list(iterable)
    assert len(items) == 1, len(items)
    return items[0]
